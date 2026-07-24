# views/energy.py
# API tính năng lượng tiêu thụ theo khoảng thời gian (giờ/tuần/tháng/năm)
#
# ĐỘNG (không hard-code meter): tự lấy tất cả meter của user + tự tìm
# register năng lượng theo từ khóa trong tên. Thêm meter mới → tự hiện.
#
# Nguyên lý: thanh ghi energy là bộ đếm TÍCH LŨY (chỉ tăng)
#   → tiêu thụ trong 1 bucket = max(value) - min(value) trong bucket đó
#   → bỏ delta âm (đồng hồ reset / tràn số)

from django.http import JsonResponse, HttpResponse
from django.utils.timezone import make_aware, localdate
from django.db.models import Count
from datetime import datetime, timedelta, time
from collections import defaultdict

from data.models import Meter, MeterRegister, Site

from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated


# ── TỪ KHÓA nhận diện register năng lượng ────────────────────────────────────
ENERGY_KEYWORDS = [
    "energy",
    "kwh",
    "active-energy",
    "real-energy",
    "forward-active",
]

# ── TỪ KHÓA loại trừ: chỉ tính năng lượng TIÊU THỤ (Forward/Import) ─────────
#   Reverse/Export = năng lượng phát ngược lưới, Reactive/Apparent = khác đơn
#   vị (kVARh/kVAh) → không được cộng chung vào kWh tiêu thụ.
DIRECTION_EXCLUDE_KEYWORDS = ["reverse", "export"]
UNIT_EXCLUDE_KEYWORDS      = ["reactive", "apparent"]

# ── FIX TẠM: các meter có CT đấu ngược chiều vật lý lúc lắp đặt ─────────────
#   → thanh ghi "Forward" của thiết bị thực ra đang đo chiều "Reverse" thật.
#   Với các meter này ta lấy Reverse làm tiêu thụ thay vì Forward.
#   TODO: gỡ khỏi set này sau khi đấu lại CT đúng chiều ngoài hiện trường.
CT_REVERSED_METER_IDS = {1}

# ── Từ khóa nhận diện pha, dùng khi meter không có thanh ghi Total riêng ────
PHASE_PATTERNS = {
    "A": ["phase-a", "-l1", "_l1", "l1-", "l1_"],
    "B": ["phase-b", "-l2", "_l2", "l2-", "l2_"],
    "C": ["phase-c", "-l3", "_l3", "l3-", "l3_"],
}

# ── Scale đổi đơn vị → kWh ───────────────────────────────────────────────────
ENERGY_SCALE = 1.0

COLOR_PALETTE = [
    "#00bcd4", "#4dd0e1", "#ffb74d", "#81c784", "#ba68c8",
    "#f06292", "#7986cb", "#a1887f", "#4db6ac", "#dce775",
]


def is_energy_register(register_name):
    if not register_name:
        return False
    name_lower = register_name.lower()
    return any(kw in name_lower for kw in ENERGY_KEYWORDS)


def get_bucket_ranges(mode, base_date):
    """Trả về list [(start, end), ...] — mỗi phần tử là 1 cột trên biểu đồ."""
    ranges = []

    if mode == "hour":
        for h in range(24):
            start = make_aware(datetime.combine(base_date, time(h, 0)))
            ranges.append((start, start + timedelta(hours=1)))

    elif mode == "week":
        monday = base_date - timedelta(days=base_date.weekday())
        for d in range(7):
            day   = monday + timedelta(days=d)
            start = make_aware(datetime.combine(day, time(0, 0)))
            ranges.append((start, start + timedelta(days=1)))

    elif mode == "month":
        first = base_date.replace(day=1)
        for w in range(4):
            start = make_aware(datetime.combine(first + timedelta(weeks=w), time(0, 0)))
            ranges.append((start, start + timedelta(weeks=1)))

    elif mode == "year":
        for m in range(1, 13):
            start = make_aware(datetime(base_date.year, m, 1))
            if m == 12:
                end = make_aware(datetime(base_date.year + 1, 1, 1))
            else:
                end = make_aware(datetime(base_date.year, m + 1, 1))
            ranges.append((start, end))

    return ranges


def select_consumption_registers(meter_id, start, end):
    """
    Chọn (các) thanh ghi đại diện cho năng lượng TIÊU THỤ (Forward/Import) của
    1 meter trong khoảng [start, end).

    Ưu tiên 1 thanh ghi "Total" duy nhất (vd Total-Forward-Active-Energy).
    Nếu meter không có Total → cộng theo pha (A/B/C), mỗi pha lấy đúng 1 tên
    thanh ghi (tên có nhiều dữ liệu nhất trong range — tránh cộng trùng khi
    thiết bị đổi tên thanh ghi giữa các thời kỳ).

    Trả về list tên thanh ghi (0, 1, hoặc tối đa 3 phần tử).
    """
    counts_qs = (MeterRegister.objects
                 .filter(meter_id=meter_id, received_at__gte=start, received_at__lt=end)
                 .values("register_name")
                 .annotate(c=Count("id")))
    counts = {row["register_name"]: row["c"] for row in counts_qs}

    # Meter có CT đấu ngược → đảo chiều Forward/Reverse khi lọc
    direction_exclude = (["forward", "import"] if meter_id in CT_REVERSED_METER_IDS
                          else DIRECTION_EXCLUDE_KEYWORDS)
    exclude_keywords = direction_exclude + UNIT_EXCLUDE_KEYWORDS

    candidates = [
        name for name in counts
        if is_energy_register(name) and not any(kw in name.lower() for kw in exclude_keywords)
    ]
    if not candidates:
        return []

    totals = [n for n in candidates if "total" in n.lower()]
    if totals:
        return [max(totals, key=lambda n: counts[n])]

    phase_best = {}
    for name in candidates:
        low = name.lower()
        for phase, patterns in PHASE_PATTERNS.items():
            if any(p in low for p in patterns):
                if phase not in phase_best or counts[name] > counts[phase_best[phase]]:
                    phase_best[phase] = name
                break
    if phase_best:
        return list(phase_best.values())

    # Không nhận diện được pha/total → lấy thanh ghi có nhiều dữ liệu nhất
    return [max(candidates, key=lambda n: counts[n])]


def get_meter_timeline(meter_id, energy_reg_names, start, end):
    """
    Trả về [(thời điểm, giá trị)] — giá trị tại mỗi thời điểm là tổng của các
    thanh ghi đã chọn (vd cộng theo pha), lấy MỘT giá trị duy nhất cho mỗi cặp
    (thanh ghi, thời điểm) — tránh bị nhân bản nếu thiết bị gửi trùng bản ghi.
    """
    if not energy_reg_names:
        return []

    qs = MeterRegister.objects.filter(
        meter_id          = meter_id,
        register_name__in = energy_reg_names,
        received_at__gte  = start,
        received_at__lt   = end,
    ).order_by("received_at")

    by_time = defaultdict(dict)
    for r in qs:
        if r.value is None:
            continue
        try:
            by_time[r.received_at][r.register_name] = float(r.value)
        except (ValueError, TypeError):
            pass

    timeline = [(t, sum(reg_vals.values())) for t, reg_vals in by_time.items()]
    return sorted(timeline)


def compute_bucket_deltas(timeline, bucket_ranges, scale):
    # timeline đã sort theo thời gian → in_bucket[0]/[-1] là mẫu đầu/cuối theo
    # THỜI GIAN thực (không phải min/max theo giá trị). Bộ đếm chỉ tăng nên
    # cách này cho kết quả đúng đồng thời không bị 1 giá trị nhiễu (đọc lỗi
    # Modbus, ...) ở giữa bucket làm phồng lên như khi dùng max(...)-min(...).
    deltas = []
    for (b_start, b_end) in bucket_ranges:
        in_bucket = [v for (t, v) in timeline if b_start <= t < b_end]
        if len(in_bucket) >= 2:
            delta = in_bucket[-1] - in_bucket[0]
            delta = max(0.0, delta) * scale
            deltas.append(round(delta, 2))
        else:
            deltas.append(0)
    return deltas


def compute_bucket_readings(timeline, bucket_ranges):
    """
    Tính chỉ số đầu kỳ, cuối kỳ và tiêu thụ (cuối - đầu) cho từng bucket, dựa
    trên mẫu đầu/cuối THEO THỜI GIAN (timeline đã sort) — bộ đếm năng lượng
    chỉ tăng nên đầu kỳ = mẫu sớm nhất, cuối kỳ = mẫu muộn nhất.
    Trả về list of (dau_ky, cuoi_ky, tieu_thu) — None khi không có dữ liệu.
    """
    result = []
    for (b_start, b_end) in bucket_ranges:
        in_bucket = [v for (t, v) in timeline if b_start <= t < b_end]
        if len(in_bucket) >= 2:
            dau_ky   = round(in_bucket[0] * ENERGY_SCALE, 2)
            cuoi_ky  = round(in_bucket[-1] * ENERGY_SCALE, 2)
            tieu_thu = round(max(0.0, cuoi_ky - dau_ky), 2)
        elif len(in_bucket) == 1:
            dau_ky   = round(in_bucket[0] * ENERGY_SCALE, 2)
            cuoi_ky  = dau_ky
            tieu_thu = 0.0
        else:
            dau_ky   = None   # không có dữ liệu → hiển thị "—"
            cuoi_ky  = None
            tieu_thu = 0.0
        result.append((dau_ky, cuoi_ky, tieu_thu))
    return result


X_LABELS = {
    "hour":  [f"{h:02d}h" for h in range(24)],
    "week":  ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
    "month": ["Week 1", "Week 2", "Week 3", "Week 4"],
    "year":  ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"],
}


# ── HÀM DÙNG CHUNG: tính series cho cả chart lẫn export ──────────────────────
def build_energy_series(user, mode, base_date, site_id=None):
    """Trả về (labels, series). Dùng chung cho energy_chart và energy_export."""
    user_site_ids = Site.objects.filter(user=user).values_list("site_id", flat=True)

    meters_qs = Meter.objects.filter(site_id__in=user_site_ids)
    if site_id:
        meters_qs = meters_qs.filter(site_id=site_id)
    meters = list(meters_qs.order_by("meter_id"))

    bucket_ranges = get_bucket_ranges(mode, base_date)
    if not bucket_ranges:
        return X_LABELS.get(mode, []), []

    range_start = bucket_ranges[0][0]
    range_end   = bucket_ranges[-1][1]

    series = []
    for idx, m in enumerate(meters):
        energy_reg_names = select_consumption_registers(m.meter_id, range_start, range_end)
        if not energy_reg_names:
            continue

        timeline = get_meter_timeline(m.meter_id, energy_reg_names, range_start, range_end)
        data = compute_bucket_deltas(timeline, bucket_ranges, ENERGY_SCALE)

        series.append({
            "key":      f"ID{m.meter_id}",
            "meter_id": m.meter_id,
            "label":    f"{m.meter_name} (ID{m.meter_id})",
            "color":    COLOR_PALETTE[idx % len(COLOR_PALETTE)],
            "data":     data,
        })

    return X_LABELS[mode], series


def parse_base_date(date_str):
    if date_str:
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return localdate()
    return localdate()


# ── API endpoint: dữ liệu biểu đồ ────────────────────────────────────────────
@api_view(["GET"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def energy_chart(request):
    mode     = request.GET.get("type", "hour")
    date_str = request.GET.get("date")
    site_id  = request.GET.get("site_id")

    if mode not in ("hour", "week", "month", "year"):
        return JsonResponse({"error": "type không hợp lệ"}, status=400)

    base_date = parse_base_date(date_str)
    labels, series = build_energy_series(request.user, mode, base_date, site_id)

    return JsonResponse({"labels": labels, "series": series})


# ── Nhãn tiêu đề khoảng thời gian (cho file Excel) ───────────────────────────
def period_title(mode, base_date):
    if mode == "hour":
        return f"Daily Report - {base_date.strftime('%d/%m/%Y')}"
    elif mode == "week":
        monday = base_date - timedelta(days=base_date.weekday())
        sunday = monday + timedelta(days=6)
        return f"Weekly Report - {monday.strftime('%d/%m/%Y')} to {sunday.strftime('%d/%m/%Y')}"
    elif mode == "month":
        return f"Monthly Report - {base_date.strftime('%m/%Y')}"
    elif mode == "year":
        return f"Yearly Report - {base_date.year}"
    return "Energy Report"


# ── API endpoint: XUẤT EXCEL ─────────────────────────────────────────────────
# Cột: Thời gian | Site | Meter | Chỉ số đầu kỳ (kWh) | Chỉ số cuối kỳ (kWh) | Tiêu thụ (kWh)
# Mỗi hàng = 1 khoảng thời gian của 1 meter
@api_view(["GET"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def energy_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    mode     = request.GET.get("type", "hour")
    date_str = request.GET.get("date")
    site_id  = request.GET.get("site_id")

    if mode not in ("hour", "week", "month", "year"):
        return JsonResponse({"error": "type không hợp lệ"}, status=400)

    base_date = parse_base_date(date_str)

    # ── Lấy danh sách meter (kèm thông tin site) ──────────────────────────
    user_site_ids = Site.objects.filter(user=request.user).values_list("site_id", flat=True)
    meters_qs = Meter.objects.filter(site_id__in=user_site_ids).select_related("site_id")
    if site_id:
        meters_qs = meters_qs.filter(site_id=site_id)
    meters = list(meters_qs.order_by("meter_id"))

    bucket_ranges = get_bucket_ranges(mode, base_date)
    if not bucket_ranges:
        return JsonResponse({"error": "Không tính được bucket"}, status=400)

    range_start = bucket_ranges[0][0]
    range_end   = bucket_ranges[-1][1]
    labels      = X_LABELS[mode]

    # Tên site cho tiêu đề
    site_name_title = "All Sites"
    if site_id:
        try:
            site_name_title = Site.objects.get(site_id=site_id, user=request.user).site_name
        except Site.DoesNotExist:
            pass

    # ── Tạo workbook ──────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Energy Report"

    # Style
    title_font  = Font(bold=True, size=13, color="FFFFFF")
    title_fill  = PatternFill("solid", fgColor="1a73e8")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="00838f")
    total_font  = Font(bold=True)
    total_fill  = PatternFill("solid", fgColor="e0f7fa")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="b0bec5")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Dòng 1: Tiêu đề report
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value=period_title(mode, base_date))
    c.font = title_font; c.fill = title_fill; c.alignment = center

    # Dòng 2: Site
    ws.merge_cells("A2:F2")
    c = ws.cell(row=2, column=1, value=f"Site: {site_name_title}")
    c.font = Font(bold=True, size=11); c.alignment = center

    # Dòng 4: Header bảng — 6 cột cố định
    headers = [
        "Time Period",
        "Site",
        "Meter",
        "Start Reading (kWh)",
        "End Reading (kWh)",
        "Consumption (kWh)",
    ]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = border

    # ── Ghi dữ liệu: mỗi meter → mỗi bucket → 1 hàng ─────────────────────
    current_row = 5
    for m in meters:
        energy_reg_names = select_consumption_registers(m.meter_id, range_start, range_end)
        if not energy_reg_names:
            continue

        timeline = get_meter_timeline(m.meter_id, energy_reg_names, range_start, range_end)
        readings = compute_bucket_readings(timeline, bucket_ranges)

        site_name_m = m.site_id.site_name
        meter_name  = f"{m.meter_name} (ID{m.meter_id})"

        total_tieu_thu = 0.0
        first_dau_ky   = None   # đầu kỳ đầu tiên trong kỳ (cho dòng tổng)
        last_cuoi_ky   = None   # cuối kỳ cuối cùng trong kỳ (cho dòng tổng)

        for i, label in enumerate(labels):
            dau_ky, cuoi_ky, tieu_thu = readings[i]

            def cell_val(v):
                return v if v is not None else "—"

            row_values = [
                label,
                site_name_m,
                meter_name,
                cell_val(dau_ky),
                cell_val(cuoi_ky),
                tieu_thu,
            ]
            for ci, val in enumerate(row_values, start=1):
                c = ws.cell(row=current_row, column=ci, value=val)
                c.alignment = center
                c.border = border

            total_tieu_thu += tieu_thu
            if first_dau_ky is None and dau_ky is not None:
                first_dau_ky = dau_ky
            if cuoi_ky is not None:
                last_cuoi_ky = cuoi_ky

            current_row += 1

        # Dòng tổng của meter này
        total_values = [
            "Period Total",
            site_name_m,
            meter_name,
            first_dau_ky if first_dau_ky is not None else "—",
            last_cuoi_ky if last_cuoi_ky  is not None else "—",
            round(total_tieu_thu, 2),
        ]
        for ci, val in enumerate(total_values, start=1):
            c = ws.cell(row=current_row, column=ci, value=val)
            c.font = total_font; c.fill = total_fill
            c.alignment = center; c.border = border

        current_row += 2   # hàng trống giữa các meter

    # Độ rộng cột
    col_widths = [14, 22, 28, 22, 22, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Trả file về trình duyệt ──
    filename = f"energy_report_{mode}_{base_date.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# Nhãn cột đầu tiên (cột thời gian) trong Excel theo mode
def X_LABELS_HEADER(mode):
    return {
        "hour":  "Hour",
        "week":  "Day",
        "month": "Week",
        "year":  "Month",
    }.get(mode, "Time")